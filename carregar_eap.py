"""
carregar_eap.py
Lê as abas CONFIG e EAP DE MEDIÇÃO do arquivo de medição de Juazeiro do Norte
e retorna dois DataFrames prontos para uso em app Streamlit.

Uso:
    from carregar_eap import carregar_dados
    config, eap = carregar_dados("EAP_MEDICAO_4_MEDICAO___OBRAS__1___2_.xlsx")
"""

import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Helpers internos
# ---------------------------------------------------------------------------

def _val(ws, row, col):
    """Lê valor de célula, retornando None se for string de erro (#REF! etc.)."""
    v = ws.cell(row, col).value
    if isinstance(v, str) and v.startswith("#"):
        return None
    return v


# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

def _extrair_config(wb) -> dict:
    """
    Extrai as informações da aba CONFIG em um dicionário estruturado.

    Blocos extraídos:
    - medicao_atual          : número da medição atual informada pelo usuário
    - vigencia               : datas e dias de vigência do contrato
    - execucao               : datas e dias do prazo de execução
    - eap_orcamento          : itens do orçamento (id, descrição, % concluído, valor R$)
    - avanco_fisico          : avanço realizado e planejado (acumulado)
    - idp                    : Índice de Desempenho de Prazo
    - tendencia              : nova duração e data de término projetada
    - cronograma_mensal      : valor planejado e medido por mês (36 meses)
    """
    ws = wb["CONFIG"]

    # ---- Medição atual ------------------------------------------------
    medicao_atual = _val(ws, 2, 2)  # célula B2

    # ---- Vigência -------------------------------------------------------
    prazo_inicio = _val(ws, 6, 12)   # L6
    prazo_dias   = _val(ws, 6, 14)   # N6
    prazo_fim    = _val(ws, 6, 15)   # O6
    vigor_decor  = _val(ws, 7, 14)   # N7  (dias decorridos)
    vigor_hoje   = _val(ws, 7, 15)   # O7  (data de referência)

    vigencia = {
        "inicio":          prazo_inicio,
        "fim":             prazo_fim,
        "prazo_total_dias": prazo_dias,
        "dias_decorridos": vigor_decor,
        "data_referencia": vigor_hoje,
        "pct_decorrido":   _val(ws, 4, 14),   # N4 (fração %)
        "pct_restante":    _val(ws, 5, 14),   # N5
    }

    # ---- Execução -------------------------------------------------------
    # Bloco execução: linhas 14-15 col L, N, O
    exec_inicio2 = _val(ws, 14, 12)  # L14 — início execução
    exec_dias2   = _val(ws, 14, 14)  # N14 — prazo total dias
    exec_fim2    = _val(ws, 14, 15)  # O14 — fim execução
    exec_decor   = _val(ws, 15, 14)  # N15 — dias decorridos
    exec_hoje    = _val(ws, 15, 15)  # O15 — data referência

    execucao = {
        "inicio":           exec_inicio2,
        "fim":              exec_fim2,
        "prazo_total_dias": exec_dias2,
        "dias_decorridos":  exec_decor,
        "data_referencia":  exec_hoje,
        "pct_decorrido":    _val(ws, 11, 14),   # N11
        "pct_restante":     _val(ws, 12, 14),   # N12
    }

    # ---- EAP Orçamento (itens de nível 1, linhas 8-17) -----------------
    eap_orcamento_rows = []
    for row in range(8, 18):
        item_id  = _val(ws, row, 1)   # col A
        descricao = _val(ws, row, 2)  # col B
        pct      = _val(ws, row, 3)   # col C
        valor    = _val(ws, row, 4)   # col D
        if descricao is not None:
            eap_orcamento_rows.append({
                "item":      item_id,
                "descricao": descricao,
                "pct_concluido": pct,
                "valor_rs":  valor,
            })

    df_eap_orc = pd.DataFrame(eap_orcamento_rows)

    # ---- Avanço físico --------------------------------------------------
    # Linha 22: avanço do período atual (medicao_atual)
    # Linha 23: avanço acumulado real (col I) e planejado (col K)
    avanco_fisico = {
        "avanco_realizado_periodo": _val(ws, 22, 9),
        "avanco_planejado_periodo": _val(ws, 22, 11),
        "avanco_realizado_acum":    _val(ws, 23, 9),
        "avanco_planejado_acum":    _val(ws, 23, 11),
    }

    # ---- IDP  (linha 27 col M) ------------------------------------------
    idp_val = _val(ws, 27, 13)
    idp = {"valor": idp_val}

    # ---- Tendência de término -------------------------------------------
    # linhas 34-37: duração original, nova duração (d), nova duração (mês), nova data
    tendencia = {
        "duracao_original_dias":  _val(ws, 34, 12),  # linha 34 col L
        "nova_duracao_dias":      _val(ws, 35, 12),
        "nova_duracao_meses":     _val(ws, 36, 12),
        "nova_data_termino":      _val(ws, 37, 12),
    }

    # ---- Cronograma mensal  ---------------------------------------------
    # Linha 46: números dos meses (cols N=14 em diante)
    # Linha 48: valor planejado por mês
    # Linha 49: valor planejado acumulado
    # Linha 50: valor medido por mês
    # Linha 51: valor medido acumulado
    PRIMEIRA_COL = 14   # col N
    NUM_MESES    = 36

    cronograma_rows = []
    for i in range(NUM_MESES):
        col = PRIMEIRA_COL + i
        mes_num = _val(ws, 46, col)
        cronograma_rows.append({
            "mes":                  int(mes_num) if isinstance(mes_num, (int, float)) else i + 1,
            "valor_planejado":      _val(ws, 48, col),
            "valor_planejado_acum": _val(ws, 49, col),
            "valor_medido":         _val(ws, 50, col),
            "valor_medido_acum":    _val(ws, 51, col),
        })

    df_cronograma = pd.DataFrame(cronograma_rows)

    return {
        "medicao_atual":   medicao_atual,
        "vigencia":        vigencia,
        "execucao":        execucao,
        "eap_orcamento":   df_eap_orc,
        "avanco_fisico":   avanco_fisico,
        "idp":             idp,
        "tendencia":       tendencia,
        "cronograma_mensal": df_cronograma,
    }


# ---------------------------------------------------------------------------
# EAP DE MEDIÇÃO
# ---------------------------------------------------------------------------

def _extrair_eap(wb) -> pd.DataFrame:
    """
    Extrai a EAP DE MEDIÇÃO num DataFrame no formato longo (tidy).

    Colunas do DataFrame:
        item, descricao, preco_total_rs,
        medicao, data_ref (primeiro dia do mês), valor_medido, pct_mes, pct_acumulado
    """
    ws = wb["EAP DE MEDIÇÃO"]

    # --- Mapeamento de medições a partir do cabeçalho ---------------
    # Linha 8: número da medição  |  Linha 7: data de referência (1º do mês)
    # Cada medição ocupa 3 colunas: Medido (R$) | % do mês | % acumulado
    # Coluna inicial da 1ª medição = 7 (col G)
    PRIMEIRA_COL_MED = 7
    COLS_POR_MED     = 3

    medicoes_map = []   # lista de dicts: {num, data, col_valor, col_pct_mes, col_pct_acum}
    col = PRIMEIRA_COL_MED
    while True:
        num_med = ws.cell(8, col).value
        if num_med is None:
            break
        if not isinstance(num_med, (int, float)):
            break   # chegou em coluna de resumo/texto
        data_ref = ws.cell(7, col).value
        medicoes_map.append({
            "num":         int(num_med),
            "data_ref":    data_ref,
            "col_valor":   col,
            "col_pct_mes": col + 1,
            "col_pct_acum": col + 2,
        })
        col += COLS_POR_MED

    # --- Itens da EAP (linhas 12 em diante) --------------------------
    LINHA_INICIO = 12
    ULTIMA_LINHA = 269   # última linha com item (item 10.1)

    # Colunas fixas: B=item, C=descrição, E=preço total
    registros = []
    for row in range(LINHA_INICIO, ULTIMA_LINHA + 1):
        item_id   = _val(ws, row, 2)    # col B
        descricao = _val(ws, row, 3)    # col C
        preco     = _val(ws, row, 5)    # col E

        if item_id is None and descricao is None:
            continue

        for m in medicoes_map:
            valor   = _val(ws, row, m["col_valor"])
            pct_mes = _val(ws, row, m["col_pct_mes"])
            pct_ac  = _val(ws, row, m["col_pct_acum"])

            registros.append({
                "item":          item_id,
                "descricao":     descricao,
                "preco_total_rs": preco,
                "medicao":       m["num"],
                "data_ref":      m["data_ref"],
                "valor_medido":  valor,
                "pct_mes":       pct_mes,
                "pct_acumulado": pct_ac,
            })

    df = pd.DataFrame(registros)

    # Converter data_ref para date
    df["data_ref"] = pd.to_datetime(df["data_ref"], errors="coerce").dt.date

    # Garantir tipos numéricos
    for col_num in ("preco_total_rs", "valor_medido", "pct_mes", "pct_acumulado"):
        df[col_num] = pd.to_numeric(df[col_num], errors="coerce")

    return df


# ---------------------------------------------------------------------------
# Linha de totais da EAP (linha 270 = linha de resumo geral)
# ---------------------------------------------------------------------------

def _extrair_totais_eap(wb) -> pd.DataFrame:
    """
    Extrai a linha de totais gerais (linha 270) da EAP para cada medição.
    Útil para o painel de resumo executivo.
    """
    ws = wb["EAP DE MEDIÇÃO"]

    LINHA_TOTAL  = 270
    PRIMEIRA_COL = 7
    COLS_POR_MED = 3

    rows = []
    col = PRIMEIRA_COL
    while True:
        num_med = ws.cell(8, col).value
        if num_med is None or not isinstance(num_med, (int, float)):
            break
        data_ref = ws.cell(7, col).value
        rows.append({
            "medicao":          int(num_med),
            "data_ref":         data_ref,
            "total_medido":     _val(ws, LINHA_TOTAL, col),
            "total_pct_mes":    _val(ws, LINHA_TOTAL, col + 1),
            "total_pct_acum":   _val(ws, LINHA_TOTAL, col + 2),
            # linha 271 = acumulado
            "total_medido_acum":  _val(ws, LINHA_TOTAL + 1, col),
        })
        col += COLS_POR_MED

    df = pd.DataFrame(rows)
    df["data_ref"] = pd.to_datetime(df["data_ref"], errors="coerce").dt.date
    for c in ("total_medido", "total_pct_mes", "total_pct_acum", "total_medido_acum"):
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


# ---------------------------------------------------------------------------
# Função pública
# ---------------------------------------------------------------------------

def carregar_dados(caminho_arquivo: str) -> tuple[dict, pd.DataFrame, pd.DataFrame]:
    """
    Carrega e consolida os dados das abas CONFIG e EAP DE MEDIÇÃO.

    Parâmetros
    ----------
    caminho_arquivo : str
        Caminho para o arquivo .xlsx.

    Retorna
    -------
    config : dict
        Dicionário com todos os parâmetros da aba CONFIG.
        Chaves principais:
            medicao_atual, vigencia, execucao, eap_orcamento (DataFrame),
            avanco_fisico, idp, tendencia, cronograma_mensal (DataFrame)

    df_eap : pd.DataFrame
        EAP no formato longo (tidy). Uma linha por (item, medição).
        Colunas: item, descricao, preco_total_rs, medicao, data_ref,
                 valor_medido, pct_mes, pct_acumulado

    df_totais : pd.DataFrame
        Linha de totais gerais por medição.
        Colunas: medicao, data_ref, total_medido, total_pct_mes,
                 total_pct_acum, total_medido_acum
    """
    caminho = Path(caminho_arquivo)
    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

    wb = load_workbook(caminho, data_only=True)

    config    = _extrair_config(wb)
    df_eap    = _extrair_eap(wb)
    df_totais = _extrair_totais_eap(wb)

    return config, df_eap, df_totais


# ---------------------------------------------------------------------------
# Execução direta — verificação rápida
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    arquivo = sys.argv[1] if len(sys.argv) > 1 else "EAP_MEDICAO_4_MEDICAO___OBRAS__1_ (2).xlsx"
    #arquivo = "C:/Users/cesar.macieira/Desktop/Trabalho/Python_312/dashboard-obras/EAP_MEDICAO_4_MEDICAO___OBRAS__1_ (2).xlsx"
    config, df_eap, df_totais = carregar_dados(arquivo)

    print("=" * 60)
    print("CONFIG — parâmetros gerais")
    print("=" * 60)
    print(f"  Medição atual          : {config['medicao_atual']}")
    print(f"  Vigência início        : {config['vigencia']['inicio']}")
    print(f"  Vigência fim           : {config['vigencia']['fim']}")
    print(f"  Dias decorridos (vig.) : {config['vigencia']['dias_decorridos']}")
    print(f"  Execução início        : {config['execucao']['inicio']}")
    print(f"  Execução fim           : {config['execucao']['fim']}")
    print(f"  IDP                    : {config['idp']['valor']}")
    print(f"  Avanço real acum.      : {config['avanco_fisico']['avanco_realizado_acum']:.2%}"
          if config['avanco_fisico']['avanco_realizado_acum'] else "  Avanço real acum.      : N/D")
    print(f"  Avanço plan. acum.     : {config['avanco_fisico']['avanco_planejado_acum']:.2%}"
          if config['avanco_fisico']['avanco_planejado_acum'] else "  Avanço plan. acum.     : N/D")
    print(f"  Tendência término      : {config['tendencia']['nova_data_termino']}")

    print()
    print("EAP Orçamento (nível 1):")
    print(config["eap_orcamento"].to_string(index=False))

    print()
    print("Cronograma mensal (primeiros 5 meses):")
    print(config["cronograma_mensal"].head().to_string(index=False))

    print()
    print("=" * 60)
    print(f"EAP DE MEDIÇÃO — {len(df_eap)} registros  |  shape: {df_eap.shape}")
    print("=" * 60)
    print(df_eap.head(10).to_string(index=False))

    print()
    print("=" * 60)
    print("TOTAIS POR MEDIÇÃO (medições com valores):")
    print("=" * 60)
    print(df_totais[df_totais["total_medido"].notna()].to_string(index=False))